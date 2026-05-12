"""Orders app — Template Views (Cart, Checkout, Order History)"""
import json
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from django.contrib.auth import get_user_model
from django.db.models import Count, DecimalField, Prefetch, Q, Sum, Value
from django.db.models.functions import Coalesce, TruncDate, TruncMonth
from django.contrib.auth.views import redirect_to_login
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils.decorators import method_decorator
from django.utils import timezone
from openpyxl import load_workbook

from apps.catalog.forms import AdminProductForm, ensure_product_placement
from apps.catalog.models import Product, ProductImage, Category, SubCategory, Brand, ProductVariant
from .models import Cart, CartItem, Order, OrderItem
from apps.notifications.services import send_order_confirmation


# ── Cart ──────────────────────────────────────────────────────────────────────
def _packet_rules(product):
    if not (product.packet_price and product.pack_quantity):
        return None
    pack_qty = int(product.pack_quantity or 0)
    if pack_qty <= 0:
        return None
    min_multiple_qty = pack_qty * 1
    return {"pack_qty": pack_qty, "min_multiple_qty": min_multiple_qty}


def _merge_product_into_cart(user, product, quantity):
    """
    Add `quantity` units of `product` into the user's cart (merges with an existing line).
    Returns (True, None) on success, (False, error_message) on validation failure.
    """
    packet_rules = _packet_rules(product)
    if not packet_rules and quantity < product.moq:
        return False, f"Minimum order quantity is {product.moq} {product.unit}."

    if packet_rules:
        pack_qty = packet_rules["pack_qty"]
        min_multiple_qty = packet_rules["min_multiple_qty"]
        if quantity % pack_qty != 0:
            return False, (
                f"This item is sold in full packets of {pack_qty}. Please use multiples of {pack_qty}."
            )
        if quantity < min_multiple_qty:
            return False, (
                f"Minimum packet order is {min_multiple_qty} {product.unit} "
                f"({min_multiple_qty // pack_qty} packet(s))."
            )

    if quantity > product.stock:
        return False, "Not enough stock available."

    cart, _ = Cart.objects.get_or_create(user=user)
    item, created = CartItem.objects.get_or_create(cart=cart, product=product)
    if not created:
        item.quantity += quantity
    else:
        item.quantity = quantity

    if packet_rules and item.quantity % packet_rules["pack_qty"] != 0:
        return False, f"This item is sold in full packets of {packet_rules['pack_qty']}."

    if item.quantity > product.stock:
        item.quantity = product.stock
    item.save()
    return True, None


def cart_view(request):
    if not request.user.is_authenticated:
        return render(
            request,
            "orders/cart.html",
            {"items": [], "total": 0, "login_prompt": True, "next_path": request.get_full_path()},
        )
    cart, _ = Cart.objects.get_or_create(user=request.user)
    items = cart.items.select_related("product").prefetch_related("product__pricing_tiers")
    for item in items:
        product = item.product
        item.packet_mode = bool(product.packet_price and product.pack_quantity)
        rules = _packet_rules(product)
        item.packet_min_qty = rules["min_multiple_qty"] if rules else product.moq
    return render(request, "orders/cart.html", {
        "cart": cart,
        "items": items,
        "total": cart.get_total(),
        "login_prompt": False,
    })


@login_required
@require_POST
def add_to_cart(request):
    data = json.loads(request.body)
    product_id = data.get("product_id")
    quantity = int(data.get("quantity", 1))

    product = get_object_or_404(Product, pk=product_id, is_active=True)

    ok, err = _merge_product_into_cart(request.user, product, quantity)
    if not ok:
        return JsonResponse({"success": False, "error": err})

    cart, _ = Cart.objects.get_or_create(user=request.user)
    cart_total = cart.get_total()
    return JsonResponse({
        "success": True,
        "message": f"Added {quantity} {product.unit} of {product.name} to cart.",
        "cart_count": cart.item_count(),
        "cart_total": float(cart_total),
    })


@login_required
@require_POST
def update_cart(request):
    data = json.loads(request.body)
    item_id = data.get("item_id")
    quantity = int(data.get("quantity", 1))

    item = get_object_or_404(CartItem, pk=item_id, cart__user=request.user)
    product = item.product

    packet_rules = _packet_rules(product)
    if not packet_rules and quantity < product.moq:
        return JsonResponse({"success": False, "error": f"Minimum order quantity is {product.moq} {product.unit}."})

    if packet_rules:
        pack_qty = packet_rules["pack_qty"]
        min_multiple_qty = packet_rules["min_multiple_qty"]
        if quantity % pack_qty != 0:
            return JsonResponse({
                "success": False,
                "error": f"This item is sold in full packets of {pack_qty}. Please use multiples of {pack_qty}.",
            })
        if quantity < min_multiple_qty:
            return JsonResponse({
                "success": False,
                "error": f"Minimum packet order is {min_multiple_qty} {product.unit} ({min_multiple_qty // pack_qty} packet(s)).",
            })

    if quantity > product.stock:
        return JsonResponse({"success": False, "error": "Not enough stock available."})

    item.quantity = quantity
    item.save()

    cart = item.cart
    return JsonResponse({
        "success": True,
        "line_total": float(item.line_total()),
        "cart_total": float(cart.get_total()),
        "unit_price": float(item.unit_price()),
    })


@login_required
@require_POST
def remove_from_cart(request):
    data = json.loads(request.body)
    item_id = data.get("item_id")
    item = get_object_or_404(CartItem, pk=item_id, cart__user=request.user)
    item.delete()
    cart = Cart.objects.get(user=request.user)
    return JsonResponse({
        "success": True,
        "cart_total": float(cart.get_total()),
        "cart_count": cart.item_count(),
    })


# ── Checkout ──────────────────────────────────────────────────────────────────

@login_required
def checkout_view(request):
    cart, _ = Cart.objects.get_or_create(user=request.user)
    items = cart.items.select_related("product").prefetch_related("product__pricing_tiers")

    if not items.exists():
        messages.error(request, "Your cart is empty.")
        return redirect("cart")

    if request.method == "POST":
        customer_name = request.POST.get("customer_name", "").strip()
        business_name = request.POST.get("business_name", "").strip()
        phone = request.POST.get("phone", "").strip()
        email = request.POST.get("email", "").strip()
        street_locality = request.POST.get("street_locality", "").strip()
        landmark = request.POST.get("landmark", "").strip()
        address = request.POST.get("address", "").strip()
        city = request.POST.get("city", "").strip()
        pincode = request.POST.get("pincode", "").strip()
        notes = request.POST.get("notes", "").strip()

        address_parts = [part for part in [street_locality, landmark, address] if part]
        full_address = ", ".join(address_parts)

        if not all([customer_name, phone, street_locality, address]):
            messages.error(request, "Please fill in all required fields.")
            return render(request, "orders/checkout.html", {
                "cart": cart,
                "items": items,
                "total": cart.get_total(),
                "user": request.user,
            })

        # Create order
        total = cart.get_total()
        order = Order.objects.create(
            user=request.user,
            customer_name=customer_name,
            business_name=business_name,
            phone=phone,
            email=email,
            address=full_address,
            city=city,
            pincode=pincode,
            notes=notes,
            payment_method="cod",
            total_amount=total,
            status="pending",
        )

        # Create order items + reduce stock
        for item in items:
            price = item.unit_price()
            OrderItem.objects.create(
                order=order,
                product=item.product,
                product_name=item.product.name,
                brand=item.product.brand,
                quantity=item.quantity,
                unit_price=price,
            )
            # Reduce stock
            product = item.product
            product.stock = max(0, product.stock - item.quantity)
            product.save()

        # Clear cart
        items.delete()

        # Send notifications
        try:
            send_order_confirmation(order)
        except Exception:
            pass  # don't block checkout on notification failure

        messages.success(request, f"Order #{order.order_number} placed successfully!")
        return redirect("order_success", order_number=order.order_number)

    return render(request, "orders/checkout.html", {
        "cart": cart,
        "items": items,
        "total": cart.get_total(),
        "user": request.user,
    })


def order_success(request, order_number):
    order = get_object_or_404(Order, order_number=order_number)
    return render(request, "orders/order_success.html", {"order": order})


@login_required
def order_quotation(request, order_number):
    if request.user.is_staff or request.user.is_admin:
        order = get_object_or_404(
            Order.objects.prefetch_related("items__product"),
            order_number=order_number,
        )
    else:
        order = get_object_or_404(
            Order.objects.prefetch_related("items__product"),
            order_number=order_number,
            user=request.user,
        )

    issued_on = timezone.localdate()
    valid_until = issued_on + timedelta(days=7)
    subtotal = sum((item.line_total() for item in order.items.all()), Decimal("0"))
    mrp_total = Decimal("0")
    discount_amount = Decimal("0")
    for item in order.items.all():
        mrp_unit = None
        if item.product and item.product.single_product_price:
            mrp_unit = Decimal(item.product.single_product_price)
        elif item.unit_price is not None:
            mrp_unit = Decimal(item.unit_price)
        if mrp_unit is None:
            continue
        line_mrp_total = mrp_unit * Decimal(item.quantity)
        line_selling_total = Decimal(item.unit_price) * Decimal(item.quantity)
        mrp_total += line_mrp_total
        discount_amount += max(Decimal("0"), line_mrp_total - line_selling_total)

    discount_pct = Decimal("0")
    if mrp_total > 0:
        discount_pct = (discount_amount * Decimal("100")) / mrp_total
    shipping = Decimal("0.00")
    download_mode = request.GET.get("download") == "1"
    return render(request, "orders/quotation.html", {
        "order": order,
        "issued_on": issued_on,
        "valid_until": valid_until,
        "subtotal": subtotal,
        "mrp_total": mrp_total,
        "discount_amount": discount_amount,
        "discount_pct": discount_pct,
        "shipping": shipping,
        "download_mode": download_mode,
    })


# ── Order History ─────────────────────────────────────────────────────────────

@login_required
def order_history(request):
    orders = Order.objects.filter(user=request.user).prefetch_related("items").order_by("-created_at")
    return render(request, "orders/order_history.html", {"orders": orders})


@login_required
def order_detail(request, order_number):
    order = get_object_or_404(Order, order_number=order_number, user=request.user)
    return render(request, "orders/order_detail.html", {"order": order})


@login_required
@require_POST
def reorder_from_order(request, order_number):
    order = get_object_or_404(
        Order.objects.prefetch_related(
            Prefetch("items", queryset=OrderItem.objects.select_related("product"))
        ),
        order_number=order_number,
        user=request.user,
    )
    added = 0
    skipped = []
    out_of_stock = []
    for oi in order.items.all():
        if not oi.product or not oi.product.is_active:
            skipped.append(f"{oi.product_name} (no longer available)")
            continue
        ok, err = _merge_product_into_cart(request.user, oi.product, oi.quantity)
        if ok:
            added += 1
        else:
            if err == "Not enough stock available.":
                out_of_stock.append(oi.product_name)
            else:
                skipped.append(f"{oi.product_name}: {err}")

    if added:
        messages.success(
            request,
            f"Added {added} line(s) from this order to your cart.",
        )
    if out_of_stock:
        messages.error(
            request,
            "Out of stock: " + ", ".join(out_of_stock),
        )
    if skipped:
        messages.warning(
            request,
            "Some items could not be added: " + "; ".join(skipped),
        )
    if not added and not skipped and not out_of_stock:
        messages.info(request, "This order has no items to reorder.")
    return redirect("cart")


# ── Admin Panel ───────────────────────────────────────────────────────────────

def _save_gallery_uploads(product, request):
    """Attach multiple extra images from the inventory form (`gallery_images` files)."""
    files = request.FILES.getlist("gallery_images")
    if not files:
        return
    start = product.gallery_images.count()
    for i, f in enumerate(files):
        ProductImage.objects.create(product=product, image=f, sort_order=start + i)


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_admin or request.user.is_staff):
            messages.error(request, "Admin access required.")
            return redirect_to_login(request.get_full_path())
        return view_func(request, *args, **kwargs)
    return wrapper


@admin_required
def admin_dashboard(request):
    source = (request.GET.get("source") or "combined").strip().lower()
    if source not in {"combined", "online", "offline"}:
        source = "combined"

    base_orders = Order.objects.all()
    if source == "offline":
        base_orders = base_orders.filter(notes__icontains="Offline order entry")
    elif source == "online":
        base_orders = base_orders.exclude(notes__icontains="Offline order entry")

    non_cancelled = base_orders.exclude(status="cancelled")
    sales_agg = non_cancelled.aggregate(total=Sum("total_amount"))
    total_sales_rs = sales_agg["total"] or 0

    unit_agg = OrderItem.objects.filter(order__in=non_cancelled).aggregate(u=Sum("quantity"))
    units_sold = unit_agg["u"] or 0

    chart_rows = list(
        Product.objects.filter(is_active=True)
        .order_by("-stock", "name")
        .values("name", "stock")[:15]
    )
    max_stock = max((r["stock"] for r in chart_rows), default=0)
    scale = max_stock if max_stock > 0 else 1
    product_stock_chart = [
        {**r, "bar_pct": round(100 * r["stock"] / scale, 1)} for r in chart_rows
    ]

    today = timezone.localdate()
    current_month_idx = today.year * 12 + (today.month - 1)
    month_starts = []
    for offset in range(11, -1, -1):
        idx = current_month_idx - offset
        year = idx // 12
        month = (idx % 12) + 1
        month_starts.append(date(year, month, 1))

    monthly_rows = (
        non_cancelled.filter(created_at__date__gte=month_starts[0])
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(total=Sum("total_amount"))
        .order_by("month")
    )
    monthly_totals_map = {
        (row["month"].year, row["month"].month): float(row["total"] or 0) for row in monthly_rows
    }
    monthly_sales_labels = [m.strftime("%b %y") for m in month_starts]
    monthly_sales_values = [
        round(monthly_totals_map.get((m.year, m.month), 0), 2) for m in month_starts
    ]

    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    weekly_rows = (
        base_orders.filter(created_at__date__gte=week_start, created_at__date__lte=week_end)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            total_sales=Sum("total_amount"),
            delivered_sales=Sum("total_amount", filter=Q(status="delivered")),
        )
        .order_by("day")
    )
    weekly_map = {row["day"]: row for row in weekly_rows}
    weekly_labels = []
    weekly_total_sales = []
    weekly_delivered_sales = []
    for i in range(7):
        current_day = week_start + timedelta(days=i)
        row = weekly_map.get(current_day)
        weekly_labels.append(current_day.strftime("%a"))
        weekly_total_sales.append(float(row["total_sales"] or 0) if row else 0)
        weekly_delivered_sales.append(float(row["delivered_sales"] or 0) if row else 0)

    total_products = Product.objects.filter(is_active=True).count()
    low_stock = Product.objects.filter(stock__lte=10, is_active=True).count()
    healthy_stock = max(total_products - low_stock, 0)
    top_demand_rows = (
        OrderItem.objects.filter(order__in=non_cancelled)
        .values("product_name")
        .annotate(total_qty=Sum("quantity"), orders_count=Count("order", distinct=True))
        .order_by("-total_qty", "-orders_count")[:5]
    )
    top_demand_products = []
    for idx, row in enumerate(top_demand_rows, start=1):
        top_demand_products.append(
            {
                "rank": idx,
                "name": row["product_name"],
                "total_qty": int(row["total_qty"] or 0),
                "orders_count": int(row["orders_count"] or 0),
            }
        )
    most_demanding_product = top_demand_products[0] if top_demand_products else None

    status_labels_lookup = dict(Order.STATUS_CHOICES)
    status_keys_in_order = [code for code, _ in Order.STATUS_CHOICES]
    status_counts_map = {
        row["status"]: row["c"]
        for row in base_orders.values("status").annotate(c=Count("id"))
    }
    order_status_labels = [status_labels_lookup[k] for k in status_keys_in_order]
    order_status_counts = [status_counts_map.get(k, 0) for k in status_keys_in_order]

    context = {
        "total_orders": base_orders.count(),
        "pending_orders": base_orders.filter(status="pending").count(),
        "total_products": total_products,
        "low_stock": low_stock,
        "healthy_stock": healthy_stock,
        "recent_orders": base_orders.order_by("-created_at")[:10],
        "selected_source": source,
        "total_sales_rs": total_sales_rs,
        "units_sold": units_sold,
        "product_stock_chart": product_stock_chart,
        "product_stock_chart_max": max_stock,
        "monthly_sales_labels": monthly_sales_labels,
        "monthly_sales_values": monthly_sales_values,
        "weekly_labels": weekly_labels,
        "weekly_total_sales": weekly_total_sales,
        "weekly_delivered_sales": weekly_delivered_sales,
        "most_demanding_product": most_demanding_product,
        "top_demand_products": top_demand_products,
        "order_status_labels": order_status_labels,
        "order_status_counts": order_status_counts,
    }
    return render(request, "admin_panel/dashboard.html", context)


@admin_required
def admin_products(request):
    products = (
        Product.objects.select_related("category", "subcategory", "brand_obj")
        .prefetch_related("pricing_tiers", "variants")
        .order_by("-updated_at")
    )
    categories = Category.objects.filter(is_active=True)
    return render(
        request,
        "admin_panel/products.html",
        {"products": products, "categories": categories, "brands": Brand.objects.filter(is_active=True)},
    )


@admin_required
@require_POST
def admin_bulk_import_products(request):
    file = request.FILES.get("import_file")
    if not file:
        messages.error(request, "Please upload an Excel (.xlsx) file.")
        return redirect("admin_products")
    if not file.name.lower().endswith(".xlsx"):
        messages.error(request, "Only .xlsx files are supported.")
        return redirect("admin_products")

    wb = load_workbook(file, data_only=True)
    ws = wb["Products"] if "Products" in wb.sheetnames else wb.active

    def normalized_headers(sheet):
        return [str(c.value).strip().lower().replace("_", " ") if c.value is not None else "" for c in sheet[1]]

    def clean_lookup_value(value):
        return str(value or "").strip().lower()

    headers = normalized_headers(ws)
    header_aliases = {
        "category": {"category"},
        "subcategory": {"subcategory", "sub category", "sub-subcategory", "subcategory / sub-subcategory"},
        "sub_subcategory": {
            "sub sub categories (optional)",
            "sub sub category (optional)",
            "sub sub categories",
            "sub sub category",
            "sub_sub_categories (optional)",
            "sub_sub_category (optional)",
        },
        "brand": {"brand"},
        "product_name": {"product name", "product"},
        "variant_name": {"variant name", "variant"},
        "price": {"price", "single s.p", "single sp", "single selling price"},
        "stock": {"stock", "stock packets"},
        "unit": {"unit"},
        "moq": {"moq"},
        "sku": {"sku"},
        "variant_sku": {"variant sku"},
        "net_quantity": {"net quantity"},
        "single_mrp": {"single mrp", "single product price", "mrp"},
        "pack_quantity": {"items qty p.packet", "items qty p packet", "packet qty", "pack quantity"},
        "packet_price": {"packet price"},
    }
    idx = {}
    for key, aliases in header_aliases.items():
        for i, header in enumerate(headers):
            if header in aliases:
                idx[key] = i
                break

    sub_subcategory_lookup = {}
    for sheet in wb.worksheets:
        if sheet.title == ws.title or sheet.title.lower() == "instructions":
            continue
        sheet_headers = normalized_headers(sheet)
        sheet_idx = {}
        for key in ("category", "subcategory", "sub_subcategory", "product_name"):
            for i, header in enumerate(sheet_headers):
                if header in header_aliases[key]:
                    sheet_idx[key] = i
                    break
        if not {"category", "subcategory", "sub_subcategory"}.issubset(sheet_idx):
            continue

        for sheet_row in sheet.iter_rows(min_row=2):
            def sheet_val(key):
                if key not in sheet_idx:
                    return ""
                cell = sheet_row[sheet_idx[key]]
                return "" if cell.value is None else str(cell.value).strip()

            sheet_category = sheet_val("category")
            sheet_subcategory = sheet_val("subcategory")
            sheet_sub_subcategory = sheet_val("sub_subcategory")
            sheet_product = sheet_val("product_name")
            if not (sheet_category and sheet_subcategory and sheet_sub_subcategory):
                continue
            sub_subcategory_lookup[
                (
                    clean_lookup_value(sheet_category),
                    clean_lookup_value(sheet_subcategory),
                    clean_lookup_value(sheet_product),
                )
            ] = sheet_sub_subcategory
    required = [
        "category",
        "subcategory",
        "brand",
        "product_name",
        "price",
        "stock",
    ]
    missing = [h for h in required if h not in headers]
    missing = [h for h in required if h not in idx]
    if missing:
        messages.error(request, f"Missing required columns: {', '.join(missing)}")
        return redirect("admin_products")

    created_products = 0
    created_variants = 0
    updated_variants = 0
    errors = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            def val(key):
                if key not in idx:
                    return ""
                cell = row[idx[key]]
                return "" if cell.value is None else str(cell.value).strip()

            category_name = val("category")
            subcategory_name = val("subcategory")
            sub_subcategory_name = val("sub_subcategory")
            brand_name = val("brand")
            product_name = val("product_name")
            if not sub_subcategory_name:
                lookup_key = (
                    clean_lookup_value(category_name),
                    clean_lookup_value(subcategory_name),
                    clean_lookup_value(product_name),
                )
                fallback_key = (
                    clean_lookup_value(category_name),
                    clean_lookup_value(subcategory_name),
                    "",
                )
                sub_subcategory_name = sub_subcategory_lookup.get(lookup_key) or sub_subcategory_lookup.get(fallback_key, "")
            net_quantity_raw = val("net_quantity")
            net_parts = net_quantity_raw.split()
            net_quantity_value = None
            net_quantity_unit = ""
            if net_parts:
                try:
                    net_quantity_value = Decimal(net_parts[0])
                    net_quantity_unit = " ".join(net_parts[1:])[:20]
                except (InvalidOperation, ValueError):
                    net_quantity_unit = net_quantity_raw[:20]
            variant_name = val("variant_name") or net_quantity_raw or "Default"
            unit = val("unit") or "pcs"
            sku = val("sku") if "sku" in idx else ""
            variant_sku = val("variant_sku") if "variant_sku" in idx else ""
            moq_raw = val("moq") or "1"
            price_raw = val("price")
            single_mrp_raw = val("single_mrp")
            pack_qty_raw = val("pack_quantity") or "1"
            packet_price_raw = val("packet_price")
            stock_raw = val("stock") or "0"

            if not (category_name and subcategory_name and brand_name and product_name and price_raw):
                errors += 1
                continue

            category, _ = Category.objects.get_or_create(name=category_name, defaults={"is_active": True})
            subcategory, _ = SubCategory.objects.get_or_create(
                category=category,
                parent=None,
                name=subcategory_name,
                defaults={"is_active": True},
            )
            if sub_subcategory_name:
                subcategory, _ = SubCategory.objects.get_or_create(
                    category=category,
                    parent=subcategory,
                    name=sub_subcategory_name,
                    defaults={"is_active": True},
                )
            brand_obj, _ = Brand.objects.get_or_create(name=brand_name, defaults={"is_active": True})

            moq = max(1, int(float(moq_raw)))
            price = Decimal(str(price_raw))
            single_mrp = Decimal(str(single_mrp_raw)) if single_mrp_raw else None
            pack_qty = max(1, int(float(pack_qty_raw)))
            packet_price = Decimal(str(packet_price_raw)) if packet_price_raw else price * Decimal(pack_qty)
            stock_packets = max(0, int(float(stock_raw)))
            product_stock = stock_packets * pack_qty

            sku_clean = (sku or "").strip()
            product = None
            created = False
            merged_row = False

            if sku_clean:
                product = Product.objects.filter(sku__iexact=sku_clean).first()
            if product is None:
                product = Product.objects.filter(
                    brand_obj=brand_obj, name__iexact=product_name.strip()
                ).first()

            if product is not None:
                merged_row = (
                    product.category_id != category.pk
                    or product.subcategory_id != subcategory.pk
                    or product.name != product_name
                )
                ensure_product_placement(
                    product,
                    category_id=category.pk,
                    subcategory_id=subcategory.pk,
                )
            else:
                product, created = Product.objects.get_or_create(
                    category=category,
                    subcategory=subcategory,
                    name=product_name,
                    defaults={
                        "brand": brand_name,
                        "brand_obj": brand_obj,
                        "unit": unit,
                        "moq": moq,
                        "stock": product_stock,
                        "sku": sku_clean or None,
                        "single_product_price": single_mrp,
                        "pack_quantity": pack_qty,
                        "packet_price": packet_price,
                        "net_quantity_value": net_quantity_value,
                        "net_quantity_unit": net_quantity_unit,
                        "is_active": True,
                    },
                )
                ensure_product_placement(product)

            if created:
                created_products += 1
            elif not merged_row:
                product.brand_obj = brand_obj
                product.brand = brand_name
                product.unit = unit or product.unit
                product.moq = moq
                if sku_clean:
                    product.sku = sku_clean
                product.subcategory = subcategory
                product.category = category
                product.single_product_price = single_mrp
                product.pack_quantity = pack_qty
                product.packet_price = packet_price
                product.net_quantity_value = net_quantity_value
                product.net_quantity_unit = net_quantity_unit
                product.stock = product_stock
                product.sync_discount_from_pack_pricing()
                product.save()

            variant_defaults = {
                "price": price,
                "stock": stock_packets,
                "is_active": True,
            }
            variant, v_created = ProductVariant.objects.get_or_create(
                product=product,
                name=variant_name,
                defaults={
                    **variant_defaults,
                    "sku": variant_sku or None,
                },
            )
            if v_created:
                created_variants += 1
            else:
                variant.price = price
                variant.stock = stock_packets
                if variant_sku:
                    variant.sku = variant_sku
                variant.size_value = net_quantity_value
                variant.size_unit = net_quantity_unit
                variant.pack_size = pack_qty
                variant.is_active = True
                variant.save()
                updated_variants += 1

            # Keep legacy product fields in sync.
            product.stock = product_stock
            product.single_product_price = single_mrp
            product.pack_quantity = pack_qty
            product.packet_price = packet_price
            product.net_quantity_value = net_quantity_value
            product.net_quantity_unit = net_quantity_unit
            product.sync_discount_from_pack_pricing()
            product.save(
                update_fields=[
                    "stock",
                    "single_product_price",
                    "pack_quantity",
                    "packet_price",
                    "net_quantity_value",
                    "net_quantity_unit",
                    "discount_percentage",
                    "updated_at",
                ]
            )

            tier = product.pricing_tiers.order_by("min_qty").first()
            if tier:
                tier.min_qty = 1
                tier.unit_price = price
                tier.max_qty = None
                tier.save()
            else:
                product.pricing_tiers.create(min_qty=1, max_qty=None, unit_price=price, label="")

        except (ValueError, TypeError, InvalidOperation):
            errors += 1
            continue

    messages.success(
        request,
        f"Import complete: {created_products} products created, "
        f"{created_variants} variants created, {updated_variants} variants updated, {errors} rows skipped.",
    )
    return redirect("admin_products")


def _decimal_from_variant(value):
    try:
        if value in (None, ""):
            return None
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _int_from_variant(value, default=0):
    try:
        if value in (None, ""):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _sync_product_from_variant_rows(product, request):
    """
    The premium add/edit form posts variant rows as JSON. Use the first filled row
    as the product's primary pack pricing, and mirror rows into ProductVariant.
    """
    raw = request.POST.get("variants_json") or "[]"
    try:
        rows = json.loads(raw)
    except json.JSONDecodeError:
        return
    if not isinstance(rows, list):
        return
    rows = [r for r in rows if isinstance(r, dict)]
    if not rows:
        return

    first = rows[0]
    single_mrp = _decimal_from_variant(first.get("single_mrp"))
    single_sp = _decimal_from_variant(first.get("single_sp"))
    packet_price = _decimal_from_variant(first.get("packet_price"))
    pack_qty = _int_from_variant(first.get("packet_qty"), default=1)
    stock_packets = _int_from_variant(first.get("stock_packets"), default=0)
    net_qty = _decimal_from_variant(first.get("net_quantity"))
    net_unit = (first.get("net_quantity_unit") or "").strip()

    if pack_qty <= 0:
        pack_qty = 1
    if single_mrp is not None:
        product.single_product_price = single_mrp
    if packet_price is not None:
        product.packet_price = packet_price
    elif single_sp is not None:
        product.packet_price = single_sp * Decimal(pack_qty)
    product.pack_quantity = pack_qty
    product.stock = max(0, stock_packets) * pack_qty
    if net_qty is not None:
        product.net_quantity_value = net_qty
    if net_unit:
        product.net_quantity_unit = net_unit
    product.sync_discount_from_pack_pricing()
    product.save()

    tier_price = single_sp
    if tier_price is None and product.packet_price and product.pack_quantity:
        tier_price = Decimal(product.packet_price) / Decimal(product.pack_quantity)
    if tier_price is not None:
        tier = product.pricing_tiers.order_by("min_qty").first()
        if tier:
            tier.min_qty = 1
            tier.unit_price = tier_price
            tier.max_qty = None
            tier.save()
        else:
            product.pricing_tiers.create(min_qty=1, max_qty=None, unit_price=tier_price, label="")

    for idx, row in enumerate(rows, start=1):
        row_pack_qty = max(1, _int_from_variant(row.get("packet_qty"), default=1))
        row_stock_packets = _int_from_variant(row.get("stock_packets"), default=0)
        row_net_qty = _decimal_from_variant(row.get("net_quantity"))
        row_price = (
            _decimal_from_variant(row.get("single_sp"))
            or _decimal_from_variant(row.get("packet_price"))
            or Decimal("0")
        )
        row_unit = (row.get("net_quantity_unit") or "").strip()
        label = []
        if row_net_qty is not None:
            label.append(str(row_net_qty.normalize()))
        if row_unit:
            label.append(row_unit)
        variant_name = " ".join(label) or f"Variant {idx}"
        variant, _ = ProductVariant.objects.get_or_create(product=product, name=variant_name)
        variant.size_value = row_net_qty
        variant.size_unit = row_unit
        variant.pack_size = row_pack_qty
        variant.price = row_price
        variant.stock = max(0, row_stock_packets)
        variant.is_active = True
        variant.save()


@admin_required
def admin_product_add(request):
    if request.method == "POST":
        form = AdminProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            _save_gallery_uploads(product, request)
            if getattr(form, "did_merge_existing", False):
                messages.success(
                    request,
                    f"Linked existing product “{product.name}” to this category path "
                    "(same SKU or name + brand). Inventory and pricing stay on the original record.",
                )
                return redirect("admin_product_edit", pk=product.pk)
            _sync_product_from_variant_rows(product, request)
            messages.success(request, f"Product “{product.name}” added.")
            return redirect("admin_products")
    else:
        form = AdminProductForm()
    return render(
        request,
        "admin_panel/product_form.html",
        {"form": form, "title": "Add product", "product": None, "gallery_images": []},
    )


@admin_required
@require_POST
def admin_gallery_image_delete(request, pk):
    img = get_object_or_404(ProductImage, pk=pk)
    product_pk = img.product_id
    if img.image:
        img.image.delete(save=False)
    img.delete()
    messages.success(request, "Extra image removed.")
    return redirect("admin_product_edit", pk=product_pk)


@admin_required
@require_POST
def admin_product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    label = str(product)
    product.delete()
    messages.success(request, f"Removed product “{label}”.")
    return redirect("admin_products")


@admin_required
def admin_product_edit(request, pk):
    product = get_object_or_404(
        Product.objects.prefetch_related("pricing_tiers"),
        pk=pk,
    )
    if request.method == "POST":
        form = AdminProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            product = form.save()
            _sync_product_from_variant_rows(product, request)
            _save_gallery_uploads(product, request)
            messages.success(request, "Product updated.")
            return redirect("admin_products")
    else:
        form = AdminProductForm(instance=product)
    gallery_images = product.gallery_images.all()
    return render(
        request,
        "admin_panel/product_form.html",
        {"form": form, "title": "Edit product", "product": product, "gallery_images": gallery_images},
    )


@admin_required
def admin_order_detail(request, order_number):
    order = get_object_or_404(Order.objects.prefetch_related("items"), order_number=order_number)
    return render(
        request,
        "admin_panel/order_detail.html",
        {"order": order, "status_choices": Order.STATUS_CHOICES},
    )


@admin_required
def admin_orders(request):
    status_filter = request.GET.get("status", "")
    orders = Order.objects.prefetch_related("items").order_by("-created_at")
    if status_filter:
        orders = orders.filter(status=status_filter)
    return render(request, "admin_panel/orders.html", {
        "orders": orders,
        "status_filter": status_filter,
        "status_choices": Order.STATUS_CHOICES,
    })


@admin_required
@require_POST
def admin_order_delete(request, order_number):
    order = get_object_or_404(Order, order_number=order_number)
    password = request.POST.get("admin_password", "")
    if not request.user.check_password(password):
        messages.error(request, "Incorrect admin password. Order was not deleted.")
        return redirect("admin_order_detail", order_number=order.order_number)

    label = order.order_number
    order.delete()
    messages.success(request, f"Order #{label} deleted permanently.")
    return redirect("admin_orders")


# ── Admin: customer dashboard ───────────────────────────────────────────────────


def _phone_key_last10(phone):
    """Normalize phone for matching guest/offline orders to an account."""
    digits = "".join(c for c in (phone or "") if c.isdigit())
    return digits[-10:] if len(digits) >= 10 else ""


# Steps shown on customer order cards (excludes cancelled).
ORDER_TIMELINE_CODES = ("pending", "confirmed", "processing", "shipped", "delivered")


def _prefetch_order_items():
    return Prefetch(
        "items",
        queryset=OrderItem.objects.select_related("product"),
    )


def _extras_by_phone_lookup():
    """Guest/offline orders keyed by last-10 digits of phone."""
    ip = _prefetch_order_items()
    unlinked = (
        Order.objects.filter(Q(user__isnull=True) | Q(notes__icontains="Offline order entry"))
        .prefetch_related(ip)
        .order_by("-created_at")
    )
    extras_by_phone = defaultdict(list)
    for order in unlinked:
        key = _phone_key_last10(order.phone)
        if key:
            extras_by_phone[key].append(order)
    return extras_by_phone


def _annotate_order_timeline(merged):
    for o in merged:
        if o.status == "cancelled":
            o.timeline_cancelled = True
        else:
            o.timeline_cancelled = False
            o.timeline_idx = (
                ORDER_TIMELINE_CODES.index(o.status)
                if o.status in ORDER_TIMELINE_CODES
                else -1
            )


def merged_orders_for_customer(user, extras_by_phone):
    """Combine account orders with phone-matched guest orders; annotate timeline fields."""
    linked = list(user.orders.all())
    key = _phone_key_last10(user.phone)
    merged = list(linked)
    seen = {o.pk for o in merged}
    if key and key in extras_by_phone:
        for o in extras_by_phone[key]:
            if o.pk not in seen:
                merged.append(o)
                seen.add(o.pk)
    merged.sort(key=lambda x: x.created_at, reverse=True)
    _annotate_order_timeline(merged)
    return merged


def _is_offline_order(order):
    return "offline order entry" in (order.notes or "").lower()


@admin_required
def admin_customer_details(request):
    """Customer accounts with orders linked by user FK and/or matching phone (guest/offline)."""
    User = get_user_model()
    item_prefetch = _prefetch_order_items()
    orders_qs = Order.objects.prefetch_related(item_prefetch).order_by("-created_at")
    zero_money = Value(0, output_field=DecimalField(max_digits=12, decimal_places=2))
    # Include staff/admin/superuser so store owners who shop still appear; skip inactive users.
    customers = list(
        User.objects.filter(is_active=True)
        .annotate(
            order_count=Count("orders", distinct=True),
            total_spent=Coalesce(Sum("orders__total_amount"), zero_money),
        )
        .prefetch_related(Prefetch("orders", queryset=orders_qs))
    )

    extras_by_phone = _extras_by_phone_lookup()

    now = timezone.now()
    new_cutoff = now - timedelta(days=14)
    active_cutoff = now - timedelta(days=30)

    for c in customers:
        merged = merged_orders_for_customer(c, extras_by_phone)
        c.merged_orders = merged
        c.display_order_count = len(merged)
        c.display_total_spent = sum((o.total_amount for o in merged), Decimal("0"))
        c.display_offline_order_count = sum(1 for o in merged if _is_offline_order(o))
        c.display_online_order_count = c.display_order_count - c.display_offline_order_count

        if c.date_joined >= new_cutoff:
            c.dashboard_segment = "new"
        elif (c.last_login and c.last_login >= active_cutoff) or (c.display_order_count or 0) > 0:
            c.dashboard_segment = "active"
        else:
            c.dashboard_segment = "idle"

    # Most recent order first (merged_orders includes linked + phone-matched guest orders).
    customers.sort(
        key=lambda c: max((o.created_at.timestamp() for o in c.merged_orders), default=-1.0),
        reverse=True,
    )

    return render(
        request,
        "admin_panel/customer_details.html",
        {
            "customers": customers,
        },
    )


@admin_required
def admin_customer_order_history(request, user_id):
    """Premium standalone page: one customer's merged order history with search/filter."""
    User = get_user_model()
    item_prefetch = _prefetch_order_items()
    orders_qs = Order.objects.prefetch_related(item_prefetch).order_by("-created_at")
    customer = get_object_or_404(
        User.objects.filter(is_active=True).prefetch_related(Prefetch("orders", queryset=orders_qs)),
        pk=user_id,
    )
    extras_by_phone = _extras_by_phone_lookup()
    merged = merged_orders_for_customer(customer, extras_by_phone)
    display_total_spent = sum((o.total_amount for o in merged), Decimal("0"))
    offline_order_count = 0
    for order in merged:
        order.order_source = "offline" if _is_offline_order(order) else "online"
        if order.order_source == "offline":
            offline_order_count += 1
    online_order_count = len(merged) - offline_order_count
    filter_chips = [
        ("all", "All"),
        ("delivered", "Delivered"),
        ("pending", "Pending"),
        ("processing", "Processing"),
    ]
    order_category_chips = [
        ("combined", f"Combined {len(merged)}"),
        ("online", f"Online {online_order_count}"),
        ("offline", f"Offline {offline_order_count}"),
    ]
    return render(
        request,
        "admin_panel/customer_order_history.html",
        {
            "customer": customer,
            "merged_orders": merged,
            "display_order_count": len(merged),
            "display_total_spent": display_total_spent,
            "online_order_count": online_order_count,
            "offline_order_count": offline_order_count,
            "status_choices": Order.STATUS_CHOICES,
            "order_timeline_codes": ORDER_TIMELINE_CODES,
            "filter_chips": filter_chips,
            "order_category_chips": order_category_chips,
        },
    )


def admin_offline_order(request):
    if request.method == "POST":
        customer_name = (request.POST.get("customer_name") or "").strip()
        phone = (request.POST.get("phone") or "").strip()
        business_name = (request.POST.get("business_name") or "").strip() or "Offline Walk-in"
        address = (request.POST.get("address") or "").strip() or "Offline Counter"
        payment_method = (request.POST.get("payment_method") or "cash").strip() or "cash"
        status = (request.POST.get("status") or "delivered").strip() or "delivered"
        email = (request.POST.get("email") or "").strip()
        city = (request.POST.get("city") or "").strip()
        pincode = (request.POST.get("pincode") or "").strip()
        notes = (request.POST.get("notes") or "").strip()
        discount_value = Decimal(request.POST.get("discount_value") or "0")
        tax_value = Decimal(request.POST.get("tax_value") or "0")
        cart_raw = request.POST.get("cart_payload") or "[]"

        if not customer_name or not phone:
            messages.error(request, "Customer name and phone are required.")
            return redirect("admin_offline_order")

        try:
            cart_items = json.loads(cart_raw)
        except json.JSONDecodeError:
            messages.error(request, "Invalid cart payload.")
            return redirect("admin_offline_order")

        if not cart_items:
            messages.error(request, "Please add at least one product.")
            return redirect("admin_offline_order")

        status_map = dict(Order.STATUS_CHOICES)
        if status not in status_map:
            status = "delivered"

        subtotal = Decimal("0")
        order_lines = []
        for item in cart_items:
            product_id = item.get("product_id")
            qty = int(item.get("quantity") or 0)
            if qty <= 0:
                continue
            product = Product.objects.filter(pk=product_id, is_active=True).first()
            if not product:
                continue
            unit_price = Decimal(str(item.get("unit_price") or 0))
            if unit_price <= 0:
                price_guess = product.get_price_for_qty(qty) or product.base_price or Decimal("0")
                unit_price = Decimal(price_guess)
            line_total = unit_price * qty
            subtotal += line_total
            order_lines.append(
                {
                    "product": product,
                    "qty": qty,
                    "unit_price": unit_price,
                }
            )

        if not order_lines:
            messages.error(request, "No valid product lines were found.")
            return redirect("admin_offline_order")

        grand_total = max(Decimal("0"), subtotal + tax_value - discount_value)
        order = Order.objects.create(
            user=request.user if request.user.is_authenticated else None,
            business_name=business_name,
            customer_name=customer_name,
            phone=phone,
            email=email,
            address=address,
            city=city,
            pincode=pincode,
            payment_method=payment_method,
            status=status,
            total_amount=grand_total,
            notes=f"Offline order entry. {notes}".strip(),
        )
        for line in order_lines:
            product = line["product"]
            qty = line["qty"]
            OrderItem.objects.create(
                order=order,
                product=product,
                product_name=product.name,
                brand=product.brand,
                quantity=qty,
                unit_price=line["unit_price"],
            )
            product.stock = max(0, int(product.stock) - qty)
            product.save(update_fields=["stock", "updated_at"])

        messages.success(request, f"Offline order #{order.order_number} saved successfully.")
        return redirect("order_success", order_number=order.order_number)

    active_products = (
        Product.objects.filter(is_active=True)
        .prefetch_related("pricing_tiers")
        .order_by("name")
    )
    product_catalog = []
    for p in active_products:
        base_price = p.base_price or p.get_price_for_qty(1) or Decimal("0")
        pack_qty = int(p.pack_quantity or 1)
        packet_price = Decimal(p.packet_price or 0)
        if packet_price <= 0:
            packet_price = Decimal(base_price or 0) * pack_qty
        product_catalog.append(
            {
                "id": p.id,
                "name": p.name,
                "brand": p.brand,
                "stock": int(p.stock or 0),
                "price": float(base_price or 0),
                "pack_qty": pack_qty,
                "packet_price": float(packet_price or 0),
                "image": p.image.url if p.image else "",
            }
        )

    context = {
        "product_catalog": product_catalog,
        "bundle_templates": [
            {"name": "Daily Grocery Kit", "items": 5, "discount": 3},
            {"name": "Family Essentials Pack", "items": 8, "discount": 5},
            {"name": "Quick Restock Combo", "items": 4, "discount": 2},
        ],
    }
    return render(request, "admin_panel/offline_order_entry.html", context)


def admin_offline_product_search(request):
    q = (request.GET.get("q") or "").strip()
    if len(q) < 1:
        return JsonResponse({"results": []})
    products = (
        Product.objects.filter(is_active=True)
        .filter(Q(name__icontains=q) | Q(brand__icontains=q))
        .prefetch_related("pricing_tiers")
        .order_by("name")[:20]
    )
    results = []
    for p in products:
        base_price = p.base_price or p.get_price_for_qty(1) or Decimal("0")
        pack_qty = int(p.pack_quantity or 1)
        packet_price = Decimal(p.packet_price or 0)
        if packet_price <= 0:
            packet_price = Decimal(base_price or 0) * pack_qty
        results.append(
            {
                "id": p.id,
                "name": p.name,
                "brand": p.brand,
                "stock": int(p.stock or 0),
                "price": float(base_price or 0),
                "pack_qty": pack_qty,
                "packet_price": float(packet_price or 0),
                "image": p.image.url if p.image else "",
            }
        )
    return JsonResponse({"results": results})


@admin_required
@require_POST
def admin_update_order_status(request):
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({"success": False, "error": "Invalid request body."})
    order_id = data.get("order_id")
    new_status = data.get("status")
    order = get_object_or_404(Order, pk=order_id)
    if new_status not in dict(Order.STATUS_CHOICES):
        return JsonResponse({"success": False, "error": "Invalid status."})
    if new_status == "delivered":
        raw = (data.get("delivery_phone") or "").strip()
        digits = "".join(c for c in raw if c.isdigit())
        if len(digits) < 10:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Enter a valid customer phone number before marking delivered (at least 10 digits).",
                    "require_phone": True,
                }
            )
        order.phone = raw if len(raw) <= 15 else digits[-10:]
    order.status = new_status
    order.save()
    return JsonResponse({"success": True})


@admin_required
@require_POST
def admin_update_stock(request):
    """Set absolute stock or apply a delta (+/-) from the inventory dashboard."""
    data = json.loads(request.body)
    product_id = data.get("product_id")
    product = get_object_or_404(Product, pk=product_id)
    if "delta" in data:
        delta = int(data["delta"])
        product.stock = max(0, int(product.stock) + delta)
    else:
        product.stock = max(0, int(data.get("stock", 0)))
    product.save(update_fields=["stock", "updated_at"])
    return JsonResponse({"success": True, "stock": product.stock})


@admin_required
@require_POST
def admin_toggle_product_active(request):
    data = json.loads(request.body)
    product = get_object_or_404(Product, pk=data.get("product_id"))
    product.is_active = bool(data.get("is_active", True))
    product.save()
    return JsonResponse({"success": True, "is_active": product.is_active})
