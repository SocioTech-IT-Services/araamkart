"""Generate product bulk import Excel templates for admin bulk import."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def main():
    root = Path(__file__).resolve().parents[1]
    outputs = [
        root / "fixtures" / "product_bulk_import_template.xlsx",
        root / "static" / "downloads" / "product_bulk_import_template.xlsx",
    ]
    for out in outputs:
        out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "Product name",
        "Brand",
        "Category",
        "Subcategory / sub-subcategory",
        "Net quantity",
        "Single MRP",
        "Single S.P",
        "Items qty P.Packet",
        "Packet price",
        "Stock packets",
    ]
    ws.append(headers)
    ws.append(
        [
            "Colgate Strong Teeth",
            "Colgate",
            "Personal care and hygiene",
            "Oral",
            "100 g",
            "80",
            "65",
            "12",
            "=G2*H2",
            "25",
        ]
    )
    for row in range(3, 102):
        ws.cell(row=row, column=9, value=f"=G{row}*H{row}")

    header_fill = PatternFill("solid", fgColor="2D5A27")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [28, 18, 26, 30, 16, 14, 14, 20, 16, 16]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    wi = wb.create_sheet("Instructions", 1)
    rows = [
        ("AaramKart — Product bulk import (.xlsx)", True),
        ("", False),
        ("How to use", True),
        ("1. Fill the Products sheet from row 2. Row 1 must keep these exact header names.", False),
        ("2. Admin → Inventory & stock → Bulk import → upload this .xlsx file.", False),
        ("", False),
        ("Required columns (all must appear in row 1)", True),
        ("Product name — Product title shown on site.", False),
        ("Brand — Brand name.", False),
        ("Category — Category name (created if missing).", False),
        ("Subcategory / sub-subcategory — root subcategory or browse path leaf.", False),
        ("Net quantity — e.g. 100 g, 500 ml, 12 pieces.", False),
        ("Single MRP — MRP of one single item.", False),
        ("Single S.P — selling price of one single item.", False),
        ("Items qty P.Packet — number of items inside one packet.", False),
        ("Packet price — auto formula: Single S.P × Items qty P.Packet.", False),
        ("Stock packets — number of packets in stock.", False),
        ("", False),
        ("Notes", True),
        ("One row = one product/primary variant. Use another row for another size.", False),
        ("Images are not imported from Excel; add on product edit after import.", False),
        ("Use the Sub Sub Categories tab only when a product belongs under a child/leaf category.", False),
        ("If Packet price formula is not calculated by Excel, the importer calculates it automatically.", False),
    ]
    for i, (text, bold) in enumerate(rows, 1):
        cell = wi.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if bold:
            cell.font = Font(bold=True)
    wi.column_dimensions["A"].width = 92

    ws_sub = wb.create_sheet("Sub Sub Categories", 1)
    ws_sub.sheet_properties.tabColor = "22C55E"
    sub_headers = [
        "Product name",
        "Category",
        "Subcategory / sub-subcategory",
        "sub_sub_categories (optional)",
    ]
    ws_sub.append(sub_headers)
    ws_sub.append(
        [
            "Colgate Strong Teeth",
            "Personal care and hygiene",
            "Oral",
            "Toothpaste",
        ]
    )
    for col in range(1, len(sub_headers) + 1):
        c = ws_sub.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, width in enumerate([28, 26, 30, 30], start=1):
        ws_sub.column_dimensions[get_column_letter(col)].width = width
    ws_sub.freeze_panes = "A2"

    for out in outputs:
        try:
            wb.save(out)
            print(f"Wrote {out}")
        except PermissionError:
            print(f"Skipped locked file: {out}")


if __name__ == "__main__":
    main()
